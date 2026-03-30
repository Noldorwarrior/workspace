#!/usr/bin/env python3
"""
verify_sums.py — Сверка сумм, проверка границ в xlsx.
Находит итоговые строки/столбцы, пересчитывает, проверяет диапазоны.

Использование:
    python verify_sums.py data.xlsx
    python verify_sums.py --json data.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)

SUM_KEYWORDS = {"итого", "всего", "сумма", "total", "sum", "итог", "σ"}
BOUNDARY_RULES = {
    "percent": {"min": 0, "max": 100, "keywords": ["%", "процент", "доля"]},
    "count": {"min": 0, "max": None, "keywords": ["кол-во", "количество", "штат", "чел"]},
}


def find_sum_rows(ws):
    """Найти строки с итогами."""
    sum_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(kw in cell.value.lower() for kw in SUM_KEYWORDS):
                    sum_rows.append(row_idx)
                    break
    return sum_rows


def check_sum_row(ws, sum_row_idx, findings):
    """Проверить итоговую строку: сумма ячеек выше должна совпадать."""
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=sum_row_idx, column=col_idx)
        if cell.value is None or not isinstance(cell.value, (int, float)):
            continue

        # Собираем числа выше до предыдущего итога или начала
        actual_sum = 0
        count = 0
        for r in range(sum_row_idx - 1, 0, -1):
            upper_cell = ws.cell(row=r, column=col_idx)
            if upper_cell.value is None:
                continue
            if isinstance(upper_cell.value, str):
                if any(kw in upper_cell.value.lower() for kw in SUM_KEYWORDS):
                    break  # предыдущий итог
                continue
            if isinstance(upper_cell.value, (int, float)):
                actual_sum += upper_cell.value
                count += 1

        if count > 1:  # есть что суммировать
            diff = abs(cell.value - actual_sum)
            if diff > 0.01:  # порог
                findings.append({
                    "severity": "error" if diff > 1 else "warning",
                    "location": f"{ws.title}!{cell.coordinate}",
                    "expected": str(actual_sum),
                    "actual": str(cell.value),
                    "description": f"Итог {cell.coordinate}={cell.value}, пересчёт={actual_sum}, разница={diff}",
                })


def check_boundaries(ws, findings):
    """Проверка граничных значений."""
    # Проверяем заголовки первой строки на ключевые слова
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        if header_cell.value and isinstance(header_cell.value, str):
            headers[col_idx] = header_cell.value.lower()

    for col_idx, header in headers.items():
        for rule_name, rule in BOUNDARY_RULES.items():
            if any(kw in header for kw in rule["keywords"]):
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if rule["min"] is not None and cell.value < rule["min"]:
                            findings.append({
                                "severity": "error",
                                "location": f"{ws.title}!{cell.coordinate}",
                                "expected": f">= {rule['min']}",
                                "actual": str(cell.value),
                                "description": f"Значение {cell.value} ниже минимума {rule['min']} ({rule_name})",
                            })
                        if rule["max"] is not None and cell.value > rule["max"]:
                            findings.append({
                                "severity": "warning",
                                "location": f"{ws.title}!{cell.coordinate}",
                                "expected": f"<= {rule['max']}",
                                "actual": str(cell.value),
                                "description": f"Значение {cell.value} выше максимума {rule['max']} ({rule_name})",
                            })


def check_negative_values(ws, findings):
    """Проверка на отрицательные значения в финансовых столбцах."""
    finance_kw = {"бюджет", "фот", "оклад", "сумма", "стоимость", "расход", "доход"}
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value and isinstance(cell.value, str):
            headers[col_idx] = cell.value.lower()

    for col_idx, header in headers.items():
        if any(kw in header for kw in finance_kw):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        findings.append({
                            "severity": "warning",
                            "location": f"{ws.title}!{cell.coordinate}",
                            "expected": ">= 0",
                            "actual": str(cell.value),
                            "description": f"Отрицательное значение в финансовом столбце: {cell.value}",
                        })


def verify(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    findings = []
    items_checked = 0

    for ws in wb.worksheets:
        sum_rows = find_sum_rows(ws)
        for sr in sum_rows:
            check_sum_row(ws, sr, findings)
            items_checked += 1

        check_boundaries(ws, findings)
        check_negative_values(ws, findings)
        # Считаем фактические числовые ячейки
        numeric_cells = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    numeric_cells += 1
        items_checked += numeric_cells

    items_warned = len([f for f in findings if f["severity"] == "warning"])
    items_failed = len([f for f in findings if f["severity"] == "error"])
    status = "fail" if items_failed > 0 else ("warn" if items_warned > 0 else "pass")

    return {
        "script": "verify_sums.py",
        "status": status,
        "details": f"Листов: {len(wb.worksheets)}, итогов проверено: {sum(len(find_sum_rows(ws)) for ws in wb.worksheets)}",
        "items_checked": items_checked,
        "items_passed": items_checked - items_warned - items_failed,
        "items_warned": items_warned,
        "items_failed": items_failed,
        "findings": findings,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("files", nargs="+")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    for f in args.files:
        if Path(f).suffix.lower() in (".xlsx", ".xls"):
            result = verify(f)
            if args.json:
                print(json.dumps(result, ensure_ascii=False, indent=2))
            else:
                print(f"📊 {f}: {result['status']}")
                for finding in result["findings"]:
                    icon = {"warning": "⚠️", "error": "❌", "info": "ℹ️"}[finding["severity"]]
                    print(f"  {icon} {finding['location']}: {finding['description']}")
